import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

# looping through spreadsheet - starting from row 2
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value # to get 'Supplier' column names
    inventory = product_list.cell(product_row, 2).value # to get 'Inventory'
    price = product_list.cell(product_row, 3).value # to get 'Price'
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    #calculation number of products per supplier
    if supplier_name in products_per_supplier: # to check if supplier name exists in dict.
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1 # add +1 count
    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # calculation inventory under 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # to add value for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

# saving to a different file
inv_file.save("inventory_with_total_value.xlsx")

